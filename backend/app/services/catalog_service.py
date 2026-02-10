from __future__ import annotations

import logging
import os
import posixpath
import re
import tempfile
import time
import hashlib
import json
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from zipfile import ZipFile

from openpyxl import load_workbook
from PIL import Image

from app.config import Settings
from app.services.graph_client import GraphClient


INVALID_DIR_CHARS = re.compile(r"[^A-Za-z0-9._-]+")
DEFAULT_SHEET_PATTERN = re.compile(r"^Sheet\d*$", re.IGNORECASE)
SAFE_CATEGORY_PATTERN = re.compile(r"^[A-Za-z0-9._-]+$")
MEDIA_CACHE_FILE_PATTERN = re.compile(r"^img_\d+\.(png|meta)$")
GRAPH_SIGNATURE_BYTES = 65536

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"

NS = {
    "main": NS_MAIN,
    "rel": NS_REL,
    "pkgrel": NS_PKG_REL,
    "xdr": NS_XDR,
    "a": NS_A,
}

logger = logging.getLogger(__name__)


@dataclass
class SheetDiagnostics:
    drawing_relationships: int = 0
    drawing_objects: int = 0
    drawing_pictures: int = 0
    embedded_image_refs: int = 0
    unknown_error_cells: int = 0
    extraction_failures: int = 0


@dataclass
class CatalogBuildResult:
    categories: list[dict[str, object]]
    workbook_source: str
    workbook_identity: str
    extraction_ms: int
    total_ms: int
    total_images: int
    category_stats: list[dict[str, object]]


@dataclass
class MediaImageResult:
    content: bytes
    etag: str
    cache_hit: bool
    workbook_source: str
    workbook_identity: str


class CatalogService:
    def __init__(self, settings: Settings, graph_client: GraphClient | None) -> None:
        self.settings = settings
        self.graph_client = graph_client
        self.cache_root = Path(tempfile.gettempdir()) / "artwork_cache"
        self.cache_root.mkdir(parents=True, exist_ok=True)
        self._identity_marker_path = self.cache_root / ".workbook_identity"
        self._last_built_workbook_identity: str | None = None

    @staticmethod
    def _workbook_identity_from_path(path: Path) -> str:
        if not path.exists() or not path.is_file():
            return f"{path}|missing"
        stat = path.stat()
        return f"{path}|{stat.st_mtime_ns}|{stat.st_size}"

    @staticmethod
    def _graph_content_signature(path: Path) -> str | None:
        if not path.exists() or not path.is_file():
            return None
        try:
            with path.open("rb") as workbook_file:
                chunk = workbook_file.read(GRAPH_SIGNATURE_BYTES)
            return hashlib.sha1(chunk).hexdigest()[:16]
        except OSError:
            return None

    def _compute_workbook_identity(self, path: Path) -> str:
        base_identity = self._workbook_identity_from_path(path)
        if self.settings.source_mode != "graph":
            return base_identity

        signature = self._graph_content_signature(path)
        if not signature:
            return base_identity
        return f"{base_identity}|sig:{signature}"

    def peek_workbook_identity(self) -> str:
        if self.settings.source_mode == "graph":
            graph_path = self._graph_tmp_workbook_path()
            return self._compute_workbook_identity(graph_path)

        local_path = self._resolve_local_xlsx_path()
        if local_path is None:
            return "local:missing"
        return self._compute_workbook_identity(local_path)

    @staticmethod
    def _safe_sheet_dir_name(sheet_name: str) -> str:
        normalized = INVALID_DIR_CHARS.sub("_", sheet_name.strip())
        return normalized or "UNTITLED"

    @staticmethod
    def _resolve_unique_dir_name(base_name: str, used_names: set[str]) -> str:
        if base_name not in used_names:
            used_names.add(base_name)
            return base_name

        suffix = 2
        while True:
            candidate = f"{base_name}_{suffix}"
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
            suffix += 1

    @staticmethod
    def _save_as_png(image_bytes: bytes, output_path: Path) -> None:
        with Image.open(BytesIO(image_bytes)) as img:
            img.save(output_path, format="PNG")

    @staticmethod
    def _to_png_bytes(image_bytes: bytes) -> bytes:
        with Image.open(BytesIO(image_bytes)) as img:
            output = BytesIO()
            img.save(output, format="PNG")
        return output.getvalue()

    @staticmethod
    def _should_ignore_sheet(sheet_name: str) -> bool:
        return bool(DEFAULT_SHEET_PATTERN.match(sheet_name.strip()))

    @staticmethod
    def _resolve_zip_target(base_part: str, target: str) -> str:
        if target.startswith("/"):
            return target.lstrip("/")
        base_dir = posixpath.dirname(base_part)
        return posixpath.normpath(posixpath.join(base_dir, target))

    @staticmethod
    def _rels_part_for(part: str) -> str:
        return posixpath.join(posixpath.dirname(part), "_rels", f"{posixpath.basename(part)}.rels")

    @staticmethod
    def _count_unknown_error_cells(worksheet) -> int:
        count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if cell.data_type == "e" and str(value).strip().upper() == "#UNKNOWN!":
                    count += 1
                elif isinstance(value, str) and value.strip().upper() == "#UNKNOWN!":
                    count += 1
        return count

    @staticmethod
    def _parse_relationships(xml_bytes: bytes) -> dict[str, tuple[str, str]]:
        root = ET.fromstring(xml_bytes)
        relationships: dict[str, tuple[str, str]] = {}
        for rel in root.findall("pkgrel:Relationship", NS):
            rel_id = rel.attrib.get("Id")
            rel_type = rel.attrib.get("Type", "")
            rel_target = rel.attrib.get("Target", "")
            if rel_id:
                relationships[rel_id] = (rel_type, rel_target)
        return relationships

    def _analyze_sheet_drawings(
        self,
        archive: ZipFile,
        sheet_part: str,
        sheet_relationships: dict[str, tuple[str, str]],
        diagnostics: SheetDiagnostics,
        mapped_media_targets: set[str],
    ) -> None:
        drawing_targets = [
            target
            for rel_type, target in sheet_relationships.values()
            if rel_type.endswith("/drawing") or "drawings/" in target
        ]
        diagnostics.drawing_relationships = len(drawing_targets)

        for drawing_target in drawing_targets:
            drawing_part = self._resolve_zip_target(sheet_part, drawing_target)
            if drawing_part not in archive.namelist():
                continue

            drawing_root = ET.fromstring(archive.read(drawing_part))
            pic_count = len(drawing_root.findall(".//xdr:pic", NS))
            sp_count = len(drawing_root.findall(".//xdr:sp", NS))
            gf_count = len(drawing_root.findall(".//xdr:graphicFrame", NS))
            grp_count = len(drawing_root.findall(".//xdr:grpSp", NS))
            cxn_count = len(drawing_root.findall(".//xdr:cxnSp", NS))
            blip_count = len(drawing_root.findall(".//a:blip", NS))
            anchor_count = (
                len(drawing_root.findall(".//xdr:oneCellAnchor", NS))
                + len(drawing_root.findall(".//xdr:twoCellAnchor", NS))
                + len(drawing_root.findall(".//xdr:absoluteAnchor", NS))
            )

            object_count = pic_count + sp_count + gf_count + grp_count + cxn_count
            if object_count == 0 and anchor_count > 0:
                object_count = anchor_count

            diagnostics.drawing_objects += object_count
            diagnostics.drawing_pictures += pic_count
            diagnostics.embedded_image_refs += blip_count

            drawing_rels_part = self._rels_part_for(drawing_part)
            if drawing_rels_part not in archive.namelist():
                continue

            drawing_relationships = self._parse_relationships(archive.read(drawing_rels_part))
            for rel_type, rel_target in drawing_relationships.values():
                if "/image" not in rel_type:
                    continue
                media_part = self._resolve_zip_target(drawing_part, rel_target)
                mapped_media_targets.add(media_part)

    def _analyze_xlsx_package(self, workbook_bytes: bytes) -> tuple[dict[str, SheetDiagnostics], int]:
        diagnostics_by_sheet: dict[str, SheetDiagnostics] = {}
        mapped_media_targets: set[str] = set()

        with ZipFile(BytesIO(workbook_bytes)) as archive:
            workbook_part = "xl/workbook.xml"
            workbook_rels_part = "xl/_rels/workbook.xml.rels"

            if workbook_part not in archive.namelist() or workbook_rels_part not in archive.namelist():
                return diagnostics_by_sheet, 0

            workbook_root = ET.fromstring(archive.read(workbook_part))
            workbook_relationships = self._parse_relationships(archive.read(workbook_rels_part))
            sheets_node = workbook_root.find("main:sheets", NS)
            if sheets_node is None:
                return diagnostics_by_sheet, 0

            for sheet in sheets_node.findall("main:sheet", NS):
                sheet_name = sheet.attrib.get("name", "")
                rel_id = sheet.attrib.get(f"{{{NS_REL}}}id")
                if not sheet_name or not rel_id or rel_id not in workbook_relationships:
                    continue

                _, sheet_target = workbook_relationships[rel_id]
                sheet_part = self._resolve_zip_target(workbook_part, sheet_target)
                diagnostics = SheetDiagnostics()
                diagnostics_by_sheet[sheet_name] = diagnostics

                sheet_rels_part = self._rels_part_for(sheet_part)
                if sheet_rels_part not in archive.namelist():
                    continue

                sheet_relationships = self._parse_relationships(archive.read(sheet_rels_part))
                self._analyze_sheet_drawings(
                    archive=archive,
                    sheet_part=sheet_part,
                    sheet_relationships=sheet_relationships,
                    diagnostics=diagnostics,
                    mapped_media_targets=mapped_media_targets,
                )

            media_parts = {
                path for path in archive.namelist() if path.startswith("xl/media/") and not path.endswith("/")
            }
            unmapped_media_count = len(media_parts - mapped_media_targets)

        return diagnostics_by_sheet, unmapped_media_count

    @staticmethod
    def _build_notes(extracted_images_count: int, diagnostics: SheetDiagnostics) -> str | None:
        if diagnostics.unknown_error_cells > 0:
            return "Worksheet contains #UNKNOWN! values; unsupported typed objects may not be extractable."

        if extracted_images_count == 0:
            if diagnostics.drawing_objects > 0 or diagnostics.embedded_image_refs > 0:
                return "Worksheet has drawing content that is not available as standard embedded images."
            return "No standard embedded images were found in this worksheet."

        unsupported_count = CatalogService._unsupported_objects_count(extracted_images_count, diagnostics)
        if unsupported_count > 0:
            return "Some worksheet objects are not standard embedded images."

        return None

    @staticmethod
    def _unsupported_objects_count(extracted_images_count: int, diagnostics: SheetDiagnostics) -> int:
        missing_extractable_pictures = max(0, diagnostics.drawing_pictures - extracted_images_count)
        non_picture_drawing_objects = max(0, diagnostics.drawing_objects - diagnostics.drawing_pictures)
        return (
            missing_extractable_pictures
            + non_picture_drawing_objects
            + diagnostics.unknown_error_cells
            + diagnostics.extraction_failures
        )

    def _resolve_local_xlsx_path(self) -> Path | None:
        if not self.settings.local_xlsx_path:
            return None

        configured_path = Path(self.settings.local_xlsx_path).expanduser()
        if configured_path.is_absolute():
            return configured_path

        return (self.settings.base_dir / configured_path).resolve()

    def _graph_tmp_workbook_path(self) -> Path:
        if os.getenv("VERCEL"):
            return Path("/tmp/artwork_graph.xlsx")
        return Path(tempfile.gettempdir()) / "artwork_graph.xlsx"

    async def _resolve_workbook_path(self) -> Path:
        if self.settings.source_mode == "graph":
            if self.graph_client is None:
                raise RuntimeError("Graph client is not configured.")
            workbook_bytes = await self.graph_client.download_excel_file()
            graph_path = self._graph_tmp_workbook_path()
            graph_path.parent.mkdir(parents=True, exist_ok=True)
            graph_path.write_bytes(workbook_bytes)
            return graph_path

        local_path = self._resolve_local_xlsx_path()
        if local_path is None:
            raise RuntimeError("No local workbook source configured. Set LOCAL_XLSX_PATH.")
        if not local_path.exists() or not local_path.is_file():
            raise FileNotFoundError(f"LOCAL_XLSX_PATH does not exist or is not a file: {local_path}")
        return local_path

    @staticmethod
    def _parse_filename_index(filename: str) -> int | None:
        match = re.fullmatch(r"img_(\d+)\.png", filename)
        if match is None:
            return None
        return int(match.group(1))

    def _resolve_sheet_by_category(self, workbook, category: str):
        used_sheet_dirs: set[str] = set()

        for worksheet in workbook.worksheets:
            if self._should_ignore_sheet(worksheet.title):
                continue

            safe_sheet_name = self._resolve_unique_dir_name(
                self._safe_sheet_dir_name(worksheet.title),
                used_sheet_dirs,
            )
            if safe_sheet_name == category:
                return worksheet

        return None

    def _cache_path(self, category: str, filename: str) -> Path:
        return self.cache_root / category / filename

    def _read_last_built_workbook_identity(self) -> str | None:
        if self._last_built_workbook_identity:
            return self._last_built_workbook_identity
        if not self._identity_marker_path.exists() or not self._identity_marker_path.is_file():
            return None
        value = self._identity_marker_path.read_text(encoding="utf-8").strip()
        if not value:
            return None
        self._last_built_workbook_identity = value
        return value

    def _store_last_built_workbook_identity(self, workbook_identity: str) -> None:
        self._last_built_workbook_identity = workbook_identity
        self._identity_marker_path.write_text(workbook_identity, encoding="utf-8")

    def _invalidate_stale_media_cache(self, old_identity: str | None, new_identity: str) -> list[str]:
        if not old_identity or old_identity == new_identity:
            return []

        cleaned_dirs: list[str] = []
        removed_png_count = 0
        removed_meta_count = 0
        for child in self.cache_root.iterdir():
            if not child.is_dir() or not SAFE_CATEGORY_PATTERN.fullmatch(child.name):
                continue

            removed_in_dir = False
            for file_path in child.iterdir():
                if not file_path.is_file():
                    continue
                if not MEDIA_CACHE_FILE_PATTERN.fullmatch(file_path.name):
                    continue
                file_path.unlink(missing_ok=True)
                removed_in_dir = True
                if file_path.suffix == ".png":
                    removed_png_count += 1
                else:
                    removed_meta_count += 1

            if removed_in_dir:
                cleaned_dirs.append(child.name)
                # Remove category dir only if no unrelated files remain.
                try:
                    next(child.iterdir())
                except StopIteration:
                    try:
                        child.rmdir()
                    except OSError:
                        pass

        logger.info(
            json.dumps(
                {
                    "event": "media_cache_invalidation",
                    "old_identity": old_identity,
                    "new_identity": new_identity,
                    "cleaned_category_dirs": cleaned_dirs if cleaned_dirs else "none",
                    "images_removed": removed_png_count,
                    "meta_removed": removed_meta_count,
                },
                separators=(",", ":"),
            )
        )
        return cleaned_dirs

    @staticmethod
    def _cache_meta_path(cache_path: Path) -> Path:
        return cache_path.with_suffix(".meta")

    @staticmethod
    def _read_cached_identity(cache_meta_path: Path) -> str | None:
        if not cache_meta_path.exists() or not cache_meta_path.is_file():
            return None
        return cache_meta_path.read_text(encoding="utf-8").strip() or None

    @staticmethod
    def _write_cached_identity(cache_meta_path: Path, workbook_identity: str) -> None:
        cache_meta_path.write_text(workbook_identity, encoding="utf-8")

    @staticmethod
    def _build_media_etag(cache_path: Path, workbook_identity: str, filename: str) -> str:
        stat = cache_path.stat()
        etag_seed = f"{workbook_identity}|{filename}"
        workbook_hash = hashlib.sha1(etag_seed.encode("utf-8")).hexdigest()[:16]
        return f'W/"{stat.st_mtime_ns:x}-{stat.st_size:x}-{workbook_hash}"'

    async def get_media_image(self, category: str, filename: str) -> MediaImageResult:
        if not SAFE_CATEGORY_PATTERN.fullmatch(category):
            raise FileNotFoundError("Invalid media category.")

        image_index = self._parse_filename_index(filename)
        if image_index is None:
            raise FileNotFoundError("Invalid media filename.")

        cache_path = self._cache_path(category, filename)
        cache_meta_path = self._cache_meta_path(cache_path)
        current_identity = self.peek_workbook_identity()

        if cache_path.exists() and cache_path.is_file():
            cached_identity = self._read_cached_identity(cache_meta_path)
            if cached_identity and cached_identity == current_identity:
                return MediaImageResult(
                    content=cache_path.read_bytes(),
                    etag=self._build_media_etag(cache_path, cached_identity, filename),
                    cache_hit=True,
                    workbook_source=cached_identity.split("|", 1)[0],
                    workbook_identity=cached_identity,
                )

        workbook_path = await self._resolve_workbook_path()
        workbook_identity = self._compute_workbook_identity(workbook_path)
        workbook = load_workbook(filename=workbook_path, data_only=True)

        try:
            worksheet = self._resolve_sheet_by_category(workbook, category)
            if worksheet is None:
                raise FileNotFoundError("Unknown media category.")

            images = getattr(worksheet, "_images", [])
            if image_index < 1 or image_index > len(images):
                raise FileNotFoundError("Image index out of range.")

            raw_bytes = images[image_index - 1]._data()
            png_bytes = self._to_png_bytes(raw_bytes)
        except FileNotFoundError:
            raise
        except Exception as exc:
            raise RuntimeError(f"Image extraction failed for {category}/{filename}: {exc}") from exc
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()

        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_bytes(png_bytes)
        self._write_cached_identity(cache_meta_path, workbook_identity)
        return MediaImageResult(
            content=png_bytes,
            etag=self._build_media_etag(cache_path, workbook_identity, filename),
            cache_hit=False,
            workbook_source=str(workbook_path),
            workbook_identity=workbook_identity,
        )

    async def build_catalog_result(self) -> CatalogBuildResult:
        started_at = time.perf_counter()
        workbook_path = await self._resolve_workbook_path()
        workbook_identity = self._compute_workbook_identity(workbook_path)
        workbook_bytes = workbook_path.read_bytes()

        extraction_started_at = time.perf_counter()
        package_diagnostics_by_sheet, unmapped_media_count = self._analyze_xlsx_package(workbook_bytes)

        self.cache_root.mkdir(parents=True, exist_ok=True)
        previous_identity = self._read_last_built_workbook_identity()
        self._invalidate_stale_media_cache(previous_identity, workbook_identity)
        self._store_last_built_workbook_identity(workbook_identity)

        workbook = load_workbook(filename=workbook_path, data_only=True)
        categories: list[dict[str, object]] = []
        category_stats: list[dict[str, object]] = []
        total_images = 0
        used_sheet_dirs: set[str] = set()

        try:
            for worksheet in workbook.worksheets:
                if self._should_ignore_sheet(worksheet.title):
                    logger.debug("Skipping default worksheet '%s' from catalog", worksheet.title)
                    continue

                safe_sheet_name = self._resolve_unique_dir_name(
                    self._safe_sheet_dir_name(worksheet.title),
                    used_sheet_dirs,
                )
                sheet_dir = self.cache_root / safe_sheet_name
                sheet_dir.mkdir(parents=True, exist_ok=True)

                diagnostics = package_diagnostics_by_sheet.get(worksheet.title, SheetDiagnostics())
                diagnostics.unknown_error_cells = self._count_unknown_error_cells(worksheet)

                images = getattr(worksheet, "_images", [])
                image_urls: list[str] = []

                for idx, image in enumerate(images, start=1):
                    try:
                        output_path = sheet_dir / f"img_{idx}.png"
                        raw_bytes = image._data()
                        self._save_as_png(raw_bytes, output_path)
                        image_urls.append(f"/api/media/{quote(safe_sheet_name)}/img_{idx}.png")
                    except Exception as exc:
                        diagnostics.extraction_failures += 1
                        logger.warning(
                            "Image extraction failed for sheet '%s' image #%s: %s",
                            worksheet.title,
                            idx,
                            exc,
                        )

                images_count = len(image_urls)
                total_images += images_count
                unsupported_count = self._unsupported_objects_count(images_count, diagnostics)
                notes = self._build_notes(images_count, diagnostics)

                logger.info(
                    "Catalog sheet '%s': extracted_images=%s drawing_objects=%s drawing_pictures=%s "
                    "unknown_error_cells=%s extraction_failures=%s unsupported_detected=%s note=%s",
                    worksheet.title,
                    images_count,
                    diagnostics.drawing_objects,
                    diagnostics.drawing_pictures,
                    diagnostics.unknown_error_cells,
                    diagnostics.extraction_failures,
                    unsupported_count > 0,
                    notes or "-",
                )

                category_stats.append(
                    {
                        "name": worksheet.title,
                        "images_count": images_count,
                        "unsupported_objects_detected": unsupported_count > 0,
                    }
                )

                categories.append(
                    {
                        "name": worksheet.title,
                        "images": image_urls,
                        "images_count": images_count,
                        "unsupported_objects_detected": unsupported_count > 0,
                        "notes": notes,
                    }
                )
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()

        if unmapped_media_count > 0:
            # ZIP fallback signal: media files exist in package without clear sheet mapping.
            logger.info(
                "Workbook contains %s media file(s) in xl/media not mapped to extracted sheet images",
                unmapped_media_count,
            )

        extraction_ms = int((time.perf_counter() - extraction_started_at) * 1000)
        total_ms = int((time.perf_counter() - started_at) * 1000)
        return CatalogBuildResult(
            categories=categories,
            workbook_source=str(workbook_path),
            workbook_identity=workbook_identity,
            extraction_ms=extraction_ms,
            total_ms=total_ms,
            total_images=total_images,
            category_stats=category_stats,
        )

    async def build_catalog(self) -> list[dict[str, object]]:
        return (await self.build_catalog_result()).categories
